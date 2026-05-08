from pathlib import Path

from openpyxl import load_workbook


def extract_xlsx(path: Path) -> str:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                rows.append("\t".join("" if c is None else str(c) for c in row))
        if rows:
            sheets.append(f"--- Sheet: {ws.title} ---\n" + "\n".join(rows))
    return "\n\n".join(sheets)
