from io import BytesIO

from openpyxl import load_workbook


def parse_xlsx(content: bytes) -> str:
    workbook = load_workbook(BytesIO(content), data_only=True)
    lines: list[str] = []

    for worksheet in workbook.worksheets:
        lines.append(f'[{worksheet.title}]')
        for row in worksheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, '')]
            if values:
                lines.append(' | '.join(values))

    return '\n'.join(lines).strip()